import openpyxl



import openpyxl

def read_excel_row(file_name: str, sheet_name: str, row_index: int):
    path = f"C:\\Users\\Monica\\PycharmProjects\\wizz_air\\test_data\\{file_name}.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]

    row_data = {}
    for j in range(1, sheet.max_column + 1):
        key = sheet.cell(row=1, column=j).value
        value = sheet.cell(row=row_index + 1, column=j).value
        row_data[key] = value
    return row_data